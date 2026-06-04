"""
CFH — calcular_iaa.py
=====================
Calcula el Inter-Annotator Agreement (IAA) entre Mireya y la segunda anotadora.

Entradas:
  data/CFH_IAA_SegundaAnotadora.xlsx  — verificación completada por segunda anotadora
  data/IAA_anotaciones_mireya.csv     — anotaciones originales de Mireya
  data/CFH_IAA_Mireya_Hibrido.xlsx   — anotaciones nuevas de Mireya (híbrido)

Salidas:
  data/CFH_IAA_Resultados.xlsx        — κ por categoría + fragmentos en desacuerdo
  data/CFH_IAA_Resultados.csv         — versión CSV para registro

Uso:
    python code/calcular_iaa.py

Autor: Mireya Camacho Celis · CFH · 2026
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# RUTAS
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_SEGUNDA  = Path("data/CFH_IAA_SegundaAnotadora.xlsx")
ANOT_MIREYA    = Path("data/IAA_anotaciones_mireya.csv")
EXCEL_MIREYA   = Path("data/CFH_IAA_Mireya_Hibrido.xlsx")
OUT_XLSX       = Path("data/CFH_IAA_Resultados.xlsx")
OUT_CSV        = Path("data/CFH_IAA_Resultados.csv")

CATEGORIAS = ["EBI", "SA", "NV", "REP"]
META_KAPPA = 0.80

# ─────────────────────────────────────────────────────────────────────────────
# PALETA
# ─────────────────────────────────────────────────────────────────────────────
AZUL_OSC = "1F4E79"; AZUL_M = "2E75B6"; AZUL_CL = "BDD7EE"
GRIS_SIST = "F2F2F2"; AMARILLO = "FFF2CC"; VERDE_CL = "E2EFDA"
VERDE_OSC = "375623"; ROJO_CL = "FCE4D6"; NARANJA = "F4B942"
BLANCO = "FFFFFF"; NEGRO = "000000"

def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def fnt(bold=False, color=NEGRO, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN κ DE COHEN
# ─────────────────────────────────────────────────────────────────────────────
def cohen_kappa(y1, y2):
    """Calcula κ de Cohen para dos listas binarias."""
    y1 = np.array(y1, dtype=int)
    y2 = np.array(y2, dtype=int)
    n = len(y1)
    if n == 0:
        return np.nan, 0

    po = np.mean(y1 == y2)  # acuerdo observado

    # Acuerdo esperado por azar
    p1_pos = np.mean(y1 == 1)
    p2_pos = np.mean(y2 == 1)
    p1_neg = 1 - p1_pos
    p2_neg = 1 - p2_pos
    pe = p1_pos * p2_pos + p1_neg * p2_neg

    if pe == 1.0:
        return 1.0 if po == 1.0 else 0.0, n

    kappa = (po - pe) / (1 - pe)
    return round(kappa, 4), n

def interpretar_kappa(k):
    if k is None or np.isnan(k): return "N/A"
    if k >= 0.80: return "✅ Sustancial-perfecto (κ≥0.80)"
    if k >= 0.60: return "🟡 Moderado-sustancial (κ≥0.60)"
    if k >= 0.40: return "🟠 Moderado (κ≥0.40)"
    return "🔴 Bajo (κ<0.40) — requiere calibración"

# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGAR EXCEL SEGUNDA ANOTADORA
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("CFH — calcular_iaa.py")
print("=" * 60)

if not EXCEL_SEGUNDA.exists():
    print(f"❌ No se encontró {EXCEL_SEGUNDA}")
    print("   Asegúrate de que la segunda anotadora haya completado el Excel.")
    sys.exit(1)

df_seg = pd.read_excel(EXCEL_SEGUNDA, sheet_name="01_Verificacion_500", header=0, skiprows=[1])
print(f"✓ Excel segunda anotadora cargado: {len(df_seg)} filas")
print(f"  Columnas: {df_seg.columns.tolist()[:10]}...")

# Renombrar columnas clave
# Estructura: ID | CORPUS | SECCIÓN | TEXTO | CONF |
#             BERT_EBI | BERT_SA | BERT_NV | BERT_REP |
#             ACUERDO_EBI | EBI_CORREG | ... | NOTAS
col_map = {}
cols = df_seg.columns.tolist()
for i, col in enumerate(cols):
    col_str = str(col).strip().upper().replace("\n", "_")
    col_map[col] = col_str
df_seg.rename(columns=col_map, inplace=True)
print(f"  Columnas renombradas: {df_seg.columns.tolist()[:12]}...")

# Detectar columnas de acuerdo y corrección
def get_col(df, candidatos):
    """Busca la primera columna que matchee algún candidato."""
    for cand in candidatos:
        matches = [c for c in df.columns if cand in str(c).upper()]
        if matches:
            return matches[0]
    return None

cols_acuerdo = {}
cols_correg  = {}
cols_bert    = {}

for cat in CATEGORIAS:
    cols_acuerdo[cat] = get_col(df_seg, [f"ACUERDO_{cat}", f"ACUERDO\n{cat}"])
    cols_correg[cat]  = get_col(df_seg, [f"{cat}_CORREG", f"{cat}\nCORREG"])
    cols_bert[cat]    = get_col(df_seg, [f"BERT_{cat}", f"BERT\n{cat}"])

print(f"\n  Columnas detectadas:")
for cat in CATEGORIAS:
    print(f"    {cat}: acuerdo={cols_acuerdo[cat]} | correg={cols_correg[cat]} | bert={cols_bert[cat]}")

# ─────────────────────────────────────────────────────────────────────────────
# 2. RECONSTRUIR ETIQUETA FINAL DE LA SEGUNDA ANOTADORA
#    Lógica: si ACUERDO=A → usar BERT; si ACUERDO=D → usar CORREG
# ─────────────────────────────────────────────────────────────────────────────
for cat in CATEGORIAS:
    col_a = cols_acuerdo[cat]
    col_c = cols_correg[cat]
    col_b = cols_bert[cat]

    if not col_a or not col_b:
        print(f"⚠  No se encontraron columnas para {cat} — omitiendo")
        df_seg[f"SEG_{cat}"] = np.nan
        continue

    def etiqueta_final(row, col_a=col_a, col_c=col_c, col_b=col_b):
        acuerdo = str(row.get(col_a, "")).strip().upper()
        if acuerdo == "A":
            try:
                return int(row[col_b])
            except (ValueError, TypeError):
                return np.nan
        elif acuerdo == "D":
            try:
                return int(row[col_c])
            except (ValueError, TypeError):
                return np.nan
        return np.nan  # no anotado aún

    df_seg[f"SEG_{cat}"] = df_seg.apply(etiqueta_final, axis=1)

# Filas con al menos una categoría anotada
df_anotado = df_seg.dropna(subset=[f"SEG_{CATEGORIAS[0]}"], how="all").copy()
n_anotado = len(df_anotado)
print(f"\n✓ Fragmentos anotados por segunda anotadora: {n_anotado} / {len(df_seg)}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. CARGAR ETIQUETAS DE MIREYA
#    Para solapamiento (84 originales) + nuevos del híbrido
# ─────────────────────────────────────────────────────────────────────────────
df_m_orig = pd.read_csv(ANOT_MIREYA)
for cat in CATEGORIAS:
    df_m_orig[cat] = df_m_orig["etiquetas_mireya"].str.contains(cat, na=False).astype(int)
df_m_orig["ID"] = df_m_orig["inner_id"].astype(str)

# Anotaciones nuevas de Mireya (híbrido) si ya las completó
df_m_nuevas = pd.DataFrame()
if EXCEL_MIREYA.exists():
    df_m_h = pd.read_excel(EXCEL_MIREYA, sheet_name="01_Nuevos_Hibrido",
                            header=0, skiprows=[1])
    # Detectar columnas TU_EBI etc.
    col_map_m = {c: str(c).strip().upper().replace("\n","_") for c in df_m_h.columns}
    df_m_h.rename(columns=col_map_m, inplace=True)
    cols_tu = {cat: get_col(df_m_h, [f"TU_{cat}"]) for cat in CATEGORIAS}
    col_id_m = get_col(df_m_h, ["ID"])

    if all(v for v in cols_tu.values()) and col_id_m:
        df_m_h["ID"] = df_m_h[col_id_m].astype(str)
        for cat in CATEGORIAS:
            df_m_h[f"M_{cat}"] = pd.to_numeric(df_m_h[cols_tu[cat]], errors="coerce")
        df_m_nuevas = df_m_h[["ID"] + [f"M_{cat}" for cat in CATEGORIAS]].dropna()
        print(f"✓ Anotaciones nuevas Mireya cargadas: {len(df_m_nuevas)} fragmentos")
    else:
        print("⚠  No se encontraron columnas TU_EBI/SA/NV/REP en el Excel Mireya")

# ─────────────────────────────────────────────────────────────────────────────
# 4. CALCULAR κ POR CATEGORÍA
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("CALCULANDO κ DE COHEN")
print("="*60)

# Merge por ID
col_id_seg = get_col(df_anotado, ["ID"])
if col_id_seg:
    df_anotado["ID"] = df_anotado[col_id_seg].astype(str)

# Combinar etiquetas de Mireya
df_m_all = df_m_orig[["ID"] + CATEGORIAS].copy()
df_m_all.columns = ["ID"] + [f"M_{cat}" for cat in CATEGORIAS]

if not df_m_nuevas.empty:
    df_m_all = pd.concat([df_m_all, df_m_nuevas]).drop_duplicates("ID")

# Join con segunda anotadora
df_join = df_anotado[["ID"] + [f"SEG_{cat}" for cat in CATEGORIAS]].merge(
    df_m_all, on="ID", how="inner"
)

print(f"Fragmentos con anotación de ambas: {len(df_join)}")

resultados = []
for cat in CATEGORIAS:
    col_m   = f"M_{cat}"
    col_seg = f"SEG_{cat}"

    if col_m not in df_join.columns or col_seg not in df_join.columns:
        resultados.append({
            "categoria": cat, "kappa": np.nan, "n": 0,
            "acuerdo_obs": np.nan, "p_m": np.nan, "p_seg": np.nan,
            "interpretacion": "N/A — datos insuficientes"
        })
        continue

    sub = df_join[[col_m, col_seg]].dropna()
    kappa, n = cohen_kappa(sub[col_m].tolist(), sub[col_seg].tolist())
    po = float(np.mean(sub[col_m].values == sub[col_seg].values))

    resultados.append({
        "categoria":    cat,
        "kappa":        kappa,
        "n":            n,
        "acuerdo_obs":  round(po, 4),
        "p_m":          round(float(sub[col_m].mean()), 3),
        "p_seg":        round(float(sub[col_seg].mean()), 3),
        "interpretacion": interpretar_kappa(kappa)
    })

    estado = "✅" if kappa >= META_KAPPA else "❌"
    print(f"\n  {cat}: κ={kappa:.4f} {estado}  ({n} pares)")
    print(f"    Acuerdo observado: {po:.1%}")
    print(f"    Prevalencia Mireya={sub[col_m].mean():.2%} | Segunda={sub[col_seg].mean():.2%}")
    print(f"    {interpretar_kappa(kappa)}")

df_res = pd.DataFrame(resultados)
kappa_promedio = df_res["kappa"].mean()
print(f"\n  κ PROMEDIO: {kappa_promedio:.4f}")
meta_alcanzada = kappa_promedio >= META_KAPPA
print(f"  META κ>0.80: {'✅ ALCANZADA' if meta_alcanzada else '❌ NO alcanzada — calibrar'}")

# ─────────────────────────────────────────────────────────────────────────────
# 5. FRAGMENTOS EN DESACUERDO (para sesión de calibración)
# ─────────────────────────────────────────────────────────────────────────────
desacuerdos = []
for cat in CATEGORIAS:
    col_m   = f"M_{cat}"
    col_seg = f"SEG_{cat}"
    if col_m not in df_join.columns or col_seg not in df_join.columns:
        continue
    sub = df_join[[col_m, col_seg, "ID"]].dropna()
    desc = sub[sub[col_m] != sub[col_seg]].copy()
    desc["categoria"] = cat
    desc.rename(columns={col_m: "mireya", col_seg: "segunda"}, inplace=True)
    desacuerdos.append(desc[["ID", "categoria", "mireya", "segunda"]])

if desacuerdos:
    df_desc = pd.concat(desacuerdos).sort_values(["ID", "categoria"])
    print(f"\n  Fragmentos en desacuerdo: {len(df_desc)}")
else:
    df_desc = pd.DataFrame(columns=["ID", "categoria", "mireya", "segunda"])

# ─────────────────────────────────────────────────────────────────────────────
# 6. GUARDAR CSV
# ─────────────────────────────────────────────────────────────────────────────
df_res.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
print(f"\n✓ Resultados CSV: {OUT_CSV}")

# ─────────────────────────────────────────────────────────────────────────────
# 7. GENERAR EXCEL DE RESULTADOS
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Hoja 1: Resumen κ ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "01_Kappa_Resumen"
ws1.column_dimensions["A"].width = 12
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 45

# Encabezado
ws1.merge_cells("A1:G1")
c = ws1.cell(row=1, column=1, value="CFH — Resultados IAA · κ de Cohen por categoría")
c.font = fnt(bold=True, color=BLANCO, size=13)
c.fill = fill(AZUL_OSC)
c.alignment = aln(h="center")
ws1.row_dimensions[1].height = 30

hdrs = ["Categoría", "κ Cohen", "n pares", "Acuerdo obs.", "Prev. Mireya", "Prev. Segunda", "Interpretación"]
for ci, h in enumerate(hdrs, 1):
    c = ws1.cell(row=2, column=ci, value=h)
    c.font = fnt(bold=True, color=BLANCO, size=11)
    c.fill = fill(AZUL_M)
    c.alignment = aln(h="center")
    c.border = borde()

for ri, row in enumerate(df_res.itertuples(), 3):
    vals = [row.categoria, row.kappa, row.n, row.acuerdo_obs,
            row.p_m, row.p_seg, row.interpretacion]
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.border = borde()
        c.alignment = aln(h="center" if ci < 7 else "left", wrap=True)
        c.font = fnt(size=11)
        # Color fila según κ
        if ci == 2 and not np.isnan(row.kappa):
            if row.kappa >= META_KAPPA:
                c.fill = fill(VERDE_CL)
            elif row.kappa >= 0.60:
                c.fill = fill(AMARILLO)
            else:
                c.fill = fill(ROJO_CL)
    ws1.row_dimensions[ri].height = 22

# Fila resumen
ri_res = len(df_res) + 3
ws1.merge_cells(f"A{ri_res}:C{ri_res}")
c = ws1.cell(row=ri_res, column=1, value=f"κ PROMEDIO: {kappa_promedio:.4f}")
c.font = fnt(bold=True, size=12,
             color=VERDE_OSC if meta_alcanzada else "CC0000")
c.fill = fill(VERDE_CL if meta_alcanzada else ROJO_CL)
c.alignment = aln(h="center")

ws1.merge_cells(f"D{ri_res}:G{ri_res}")
estado_txt = "✅ META ALCANZADA — CFH-BERT v3 desbloqueado" if meta_alcanzada else \
             "❌ META NO ALCANZADA — realizar sesión de calibración antes de continuar"
c = ws1.cell(row=ri_res, column=4, value=estado_txt)
c.font = fnt(bold=True, size=11,
             color=VERDE_OSC if meta_alcanzada else "CC0000")
c.fill = fill(VERDE_CL if meta_alcanzada else ROJO_CL)
c.alignment = aln(h="left")
ws1.row_dimensions[ri_res].height = 25

# ── Hoja 2: Desacuerdos para calibración ────────────────────────────────────
ws2 = wb.create_sheet("02_Desacuerdos_Calibracion")
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 80

# Encabezado
ws2.merge_cells("A1:E1")
c = ws2.cell(row=1, column=1,
             value="Fragmentos en desacuerdo — revisar juntas en sesión de calibración")
c.font = fnt(bold=True, color=BLANCO, size=12)
c.fill = fill(AZUL_M)
c.alignment = aln(h="center")

for ci, h in enumerate(["ID", "Categoría", "Mireya", "Segunda", "Acción sugerida"], 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = fnt(bold=True, color=BLANCO)
    c.fill = fill(NARANJA)
    c.alignment = aln(h="center")
    c.border = borde()

for ri, row in enumerate(df_desc.itertuples(), 3):
    accion = (f"Revisar definición de {row.categoria} — "
              f"Mireya={int(row.mireya)} vs Segunda={int(row.segunda)}")
    for ci, val in enumerate([row.ID, row.categoria,
                               int(row.mireya), int(row.segunda), accion], 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.border = borde()
        c.alignment = aln(wrap=True)
        c.font = fnt(size=10)
        c.fill = fill(ROJO_CL if ci <= 4 else "FFFFFF")
    ws2.row_dimensions[ri].height = 20

if len(df_desc) == 0:
    ws2.cell(row=3, column=1, value="✅ Sin desacuerdos — κ perfecto").font = fnt(bold=True, color=VERDE_OSC)

# ── Hoja 3: Protocolo si κ < 0.80 ───────────────────────────────────────────
ws3 = wb.create_sheet("03_Protocolo_Calibracion")
ws3.column_dimensions["A"].width = 95

def w3(row, texto, bold=False, color=NEGRO, bg=None, size=11):
    c = ws3.cell(row=row, column=1, value=texto)
    c.font = fnt(bold=bold, color=color, size=size)
    c.alignment = aln(wrap=True)
    ws3.row_dimensions[row].height = 18
    if bg: c.fill = fill(bg)

r = 1
w3(r, "PROTOCOLO DE CALIBRACIÓN — κ < 0.80", bold=True, color=BLANCO, bg=AZUL_OSC, size=13); r+=1
w3(r, "Pasos a seguir si alguna categoría no alcanza la meta:", italic=False); r+=2

pasos = [
    ("PASO 1 — Revisar los desacuerdos juntas (hoja 02_Desacuerdos)",
     "Repasar cada fragmento en desacuerdo. Discutir por qué cada una anotó diferente. "
     "No buscar que una 'gane' — buscar aclarar la definición."),
    ("PASO 2 — Actualizar el protocolo de anotación",
     "Si el desacuerdo revela ambigüedad en la definición (ej. frontera EBI vs NV), "
     "actualizar los criterios en el Excel de instrucciones con ejemplos nuevos."),
    ("PASO 3 — Re-anotar los fragmentos en desacuerdo",
     "Después de la calibración, cada anotadora re-anota independientemente los "
     "fragmentos problemáticos. Recalcular κ con calcular_iaa.py."),
    ("PASO 4 — Si κ persiste < 0.60 en una categoría",
     "Evaluar si la categoría necesita subdivisión o si los criterios son "
     "estructuralmente ambiguos. Documentar la decisión en el capítulo de metodología."),
]

for titulo, desc in pasos:
    w3(r, titulo, bold=True, bg=AZUL_CL, color=AZUL_OSC); r+=1
    w3(r, f"  {desc}"); r+=2

w3(r, "UMBRALES DE REFERENCIA (Landis & Koch, 1977)", bold=True, color=BLANCO, bg=AZUL_M); r+=1
for rango, label in [
    ("κ ≥ 0.80", "Sustancial a perfecto — META CFH"),
    ("0.60 ≤ κ < 0.80", "Moderado a sustancial — aceptable con calibración"),
    ("0.40 ≤ κ < 0.60", "Moderado — requiere calibración obligatoria"),
    ("κ < 0.40", "Bajo — redefinir la categoría antes de continuar"),
]:
    w3(r, f"  {rango}:  {label}"); r+=1

# ─────────────────────────────────────────────────────────────────────────────
wb.save(OUT_XLSX)
print(f"✓ Excel resultados: {OUT_XLSX}")

print(f"""
{"="*60}
RESUMEN IAA
{"="*60}
Fragmentos comparados : {len(df_join)}
κ promedio            : {kappa_promedio:.4f}
Meta κ > 0.80         : {"✅ ALCANZADA" if meta_alcanzada else "❌ NO alcanzada"}
Desacuerdos totales   : {len(df_desc)}

Archivos:
  {OUT_XLSX}
  {OUT_CSV}
{"="*60}
""")

if not meta_alcanzada:
    cats_bajas = df_res[df_res["kappa"] < META_KAPPA]["categoria"].tolist()
    print(f"Categorías bajo la meta: {cats_bajas}")
    print("→ Ver hoja 02_Desacuerdos_Calibracion para los fragmentos a revisar")
    print("→ Ver hoja 03_Protocolo_Calibracion para el protocolo de calibración")

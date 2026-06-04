"""
CFH — limpiar_excel_ciego.py
=============================
Elimina filas sin texto del CFH_IAA_Ciego_84.xlsx
y genera CFH_IAA_Ciego_60.xlsx listo para enviar.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

IN_PATH  = Path("data/CFH_IAA_Ciego_84.xlsx")
OUT_PATH = Path("data/CFH_IAA_Ciego_60.xlsx")

AZUL_OSC = "1F4E79"; AZUL_M = "2E75B6"; GRIS_SIST = "F2F2F2"
AMARILLO = "FFF2CC"; NARANJA = "F4B942"; BLANCO = "FFFFFF"; NEGRO = "000000"

def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def fnt(bold=False, color=NEGRO, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

print("=" * 50)
print("CFH — limpiar_excel_ciego.py")
print("=" * 50)

# Cargar workbook original para copiar hojas 00 y 01
wb_orig = load_workbook(IN_PATH)
wb_new = Workbook()
wb_new.remove(wb_new.active)

for sheet_name in ["00_Instrucciones", "01_Calibracion"]:
    if sheet_name in wb_orig.sheetnames:
        ws_o = wb_orig[sheet_name]
        ws_n = wb_new.create_sheet(sheet_name)
        for row in ws_o.iter_rows():
            for cell in row:
                nc = ws_n.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    nc.font      = cell.font.copy()
                    nc.fill      = cell.fill.copy()
                    nc.alignment = cell.alignment.copy()
                    nc.border    = cell.border.copy()
        for col, dim in ws_o.column_dimensions.items():
            ws_n.column_dimensions[col].width = dim.width
        for row, dim in ws_o.row_dimensions.items():
            ws_n.row_dimensions[row].height = dim.height
        for merged in ws_o.merged_cells.ranges:
            ws_n.merge_cells(str(merged))

# Leer datos y filtrar sin texto
df = pd.read_excel(IN_PATH, sheet_name="02_Anotacion", skiprows=[1])
col_texto = df.columns[3]
mask = ~(df[col_texto].isna() | (df[col_texto].astype(str).str.strip() == ''))
df_clean = df[mask].reset_index(drop=True)

print(f"Originales : 84")
print(f"Con texto  : {len(df_clean)}")
print(f"Eliminados : {84 - len(df_clean)}")

# Hoja 02 limpia
ws2 = wb_new.create_sheet("02_Anotacion")
hdrs = [
    ("ID",10,AZUL_M,BLANCO),("CORPUS",13,AZUL_M,BLANCO),
    ("SECCIÓN",16,AZUL_M,BLANCO),("TEXTO",80,AZUL_M,BLANCO),
    ("EBI",10,NARANJA,NEGRO),("SA",10,NARANJA,NEGRO),
    ("NV",10,NARANJA,NEGRO),("REP",10,NARANJA,NEGRO),
    ("NOTAS",40,AMARILLO,NEGRO),
]
for ci,(hdr,ancho,bg,fg) in enumerate(hdrs,1):
    c = ws2.cell(row=1,column=ci,value=hdr)
    c.font = fnt(bold=True,color=fg,size=11)
    c.fill = fill(bg)
    c.alignment = aln(h="center")
    c.border = borde()
    ws2.column_dimensions[get_column_letter(ci)].width = ancho
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:D2")
ws2.cell(row=2,column=1,value="◄ SISTEMA — no modificar").font = fnt(italic=True,color="595959",size=9)
ws2.merge_cells("E2:I2")
c = ws2.cell(row=2,column=5,
    value="◄ TU ANOTACIÓN — escribe 0 (ausente) o 1 (presente). Duda → pon 0 y escribe en NOTAS.")
c.font = fnt(bold=True,color="7F4800",size=10)
c.fill = fill(AMARILLO)
c.alignment = aln(wrap=True)

dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
dv.error = "Solo 0 o 1"
dv.prompt = "0 = Ausente  ·  1 = Presente"
ws2.add_data_validation(dv)

for ri,(_,row) in enumerate(df_clean.iterrows(),3):
    texto = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
    if len(texto) > 1000:
        texto = texto[:1000] + "..."
    vals = [str(row.iloc[0]),str(row.iloc[1]),str(row.iloc[2]),
            texto,"","","","",""]
    for ci,val in enumerate(vals,1):
        c = ws2.cell(row=ri,column=ci,value=val)
        c.alignment = aln(wrap=True)
        c.border = borde()
        c.font = fnt(size=10)
        c.fill = fill(GRIS_SIST if ci<=4 else AMARILLO)
    ws2.row_dimensions[ri].height = 60
    for col_bin in [5,6,7,8]:
        dv.add(f"{get_column_letter(col_bin)}{ri}")

ws2.freeze_panes = "E3"
wb_new.save(OUT_PATH)
print(f"\n✓ Guardado: {OUT_PATH}")
print("  Listo para enviar a la segunda anotadora.")

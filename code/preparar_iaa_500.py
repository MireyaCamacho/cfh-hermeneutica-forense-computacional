"""
CFH — preparar_iaa_500.py  (v2 — lectura real del corpus)
==========================================================
Lee texto real desde data/processed/corpus_a y corpus_b,
pre-anota con CFH-BERT v2, y genera dos Excels:

  data/CFH_IAA_SegundaAnotadora.xlsx  — verificación (500 fragmentos con etiquetas BERT)
  data/CFH_IAA_Mireya_Hibrido.xlsx    — híbrido (410 nuevos, etiquetas BERT ocultas)

Uso:
    python code/preparar_iaa_500.py

Autor: Mireya Camacho Celis · CFH · 2026
"""

import json
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import torch
    from transformers import AutoModelForSequenceClassification, AutoTokenizer
    TORCH_OK = True
except ImportError:
    TORCH_OK = False
    print("⚠  torch/transformers no instalado — se usará pre-anotación de demostración.")

# ─────────────────────────────────────────────────────────────────────────────
# RUTAS — ajusta si difieren
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH        = Path("cfh.db")
MODELO_PATH    = Path("models")
CORPUS_A_PATH  = Path("data/processed/corpus_a")
CORPUS_B_PATH  = Path("data/processed/corpus_b")
ANOT_MIREYA    = Path("data/IAA_anotaciones_mireya.csv")
OUT_SEGUNDA    = Path("data/CFH_IAA_SegundaAnotadora.xlsx")
OUT_MIREYA     = Path("data/CFH_IAA_Mireya_Hibrido.xlsx")

N_TOTAL     = 500
N_MIREYA    = 84    # anotaciones reales existentes
N_NUEVOS    = N_TOTAL - N_MIREYA   # 416
SEED        = 42
CATEGORIAS  = ["EBI", "SA", "NV", "REP"]

# ─────────────────────────────────────────────────────────────────────────────
# PALETA
# ─────────────────────────────────────────────────────────────────────────────
AZUL_OSC = "1F4E79"; AZUL_M = "2E75B6"; AZUL_CL = "BDD7EE"
GRIS_SIST = "F2F2F2"; GRIS_H = "D9D9D9"; AMARILLO = "FFF2CC"
VERDE_CL = "E2EFDA"; VERDE_OSC = "375623"; LILA_CL = "EAD1DC"
NARANJA = "F4B942"; BLANCO = "FFFFFF"; NEGRO = "000000"

def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def fnt(bold=False, color=NEGRO, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# 1. ÍNDICE DE TEXTOS: sha256 → ruta .txt
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("CFH — preparar_iaa_500.py v2")
print("=" * 60)

def construir_indice_textos():
    """Construye dict sha256_clean → Path(.txt) leyendo los JSON de corpus A y B."""
    indice = {}
    for corpus_path in [CORPUS_A_PATH, CORPUS_B_PATH]:
        if not corpus_path.exists():
            continue
        for json_path in corpus_path.glob("*.json"):
            try:
                with open(json_path, encoding="utf-8") as f:
                    meta = json.load(f)
                sha = meta.get("sha256_clean", "")
                txt_path = json_path.with_suffix(".txt")
                if sha and txt_path.exists():
                    indice[sha] = (txt_path, meta)
                # También indexar por nombre de archivo (sin extensión)
                indice[json_path.stem] = (txt_path, meta)
            except Exception:
                pass
    print(f"✓ Índice de textos: {len(indice)} documentos")
    return indice

indice_textos = construir_indice_textos()

# ─────────────────────────────────────────────────────────────────────────────
# 2. CARGAR BLOQUES CON TEXTO REAL DESDE DB + JSON/TXT
# ─────────────────────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)

# Cargar bloques con info de documento
df_bloques = pd.read_sql("""
    SELECT b.id, b.seccion, b.orden, b.n_chars,
           b.identificador_externo,
           d.id as doc_id, d.corpus AS corpus_id
    FROM bloques b
    JOIN documentos d ON b.documento_id = d.id
    WHERE b.seccion NOT IN ('ENCABEZADO', 'PORTADA')
      AND b.n_chars >= 100
""", conn)

print(f"✓ Bloques cargados desde DB: {len(df_bloques)}")
print(f"  Corpus: {df_bloques['corpus_id'].value_counts().to_dict()}")

# Función para extraer texto de un bloque desde el JSON/TXT
def extraer_texto_bloque(row, indice):
    """Busca el txt del documento y extrae el fragmento por posición de sección."""
    id_ext = str(row.get("identificador_externo", ""))
    # identificador_externo = SHA256_64chars__SECCION__orden
    # El JSON se nombra con los primeros 16 chars del SHA256
    sha_full = id_ext.split("__")[0] if "__" in id_ext else id_ext
    sha_stem = sha_full[:16]

    entrada = indice.get(sha_stem) or indice.get(sha_full) or indice.get(id_ext)
    if not entrada:
        return ""

    txt_path, meta = entrada
    try:
        txt = txt_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""

    # Encontrar la sección en el JSON y extraer por char_range
    seccion_buscada = row["seccion"]
    orden = int(row.get("orden", 1))
    secciones = meta.get("segmentation", {}).get("sections", [])
    for sec in secciones:
        if sec.get("section_id") == seccion_buscada:
            cr = sec.get("char_range", [])
            if len(cr) == 2:
                fragmento = txt[cr[0]:cr[1]]
                palabras = fragmento.split()
                # Dividir en bloques de ~200 palabras según el orden del bloque
                bloque_size = 200
                inicio = (orden - 1) * bloque_size
                fin = min(inicio + bloque_size, len(palabras))
                return " ".join(palabras[inicio:fin])
    return ""

# Extraer texto para muestra primero (para verificar)
muestra = df_bloques.head(3).copy()
muestra["texto_real"] = muestra.apply(lambda r: extraer_texto_bloque(r, indice_textos), axis=1)
print("\nVerificación de extracción de texto (3 bloques):")
for _, row in muestra.iterrows():
    print(f"  Bloque {row['id']} | {row['corpus_id']} | {row['seccion']}")
    print(f"  Texto: {row['texto_real'][:120]}...")
    print()

conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# 3. CARGAR ANOTACIONES DE MIREYA Y PARSEAR ETIQUETAS
# ─────────────────────────────────────────────────────────────────────────────
df_anot = pd.read_csv(ANOT_MIREYA)

# Parsear etiquetas_combinadas → columnas binarias
for cat in CATEGORIAS:
    df_anot[cat] = df_anot["etiquetas_mireya"].str.contains(cat, na=False).astype(int)

# IDs anotados = inner_id (que corresponde a bloques.id)
ids_anotados = set(df_anot["inner_id"].unique())
print(f"✓ Anotaciones Mireya cargadas: {len(df_anot)} fragmentos")
print(f"  EBI={df_anot['EBI'].sum()} SA={df_anot['SA'].sum()} "
      f"NV={df_anot['NV'].sum()} REP={df_anot['REP'].sum()}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. CONSTRUIR DATASET DE 500 FRAGMENTOS
# ─────────────────────────────────────────────────────────────────────────────

# Dataset Mireya: bloques que ella anotó con texto real
conn = sqlite3.connect(DB_PATH)
df_mireya_bloques = pd.read_sql(f"""
    SELECT b.id, b.seccion, b.orden, b.n_chars,
           d.corpus AS corpus_id, b.identificador_externo
    FROM bloques b
    JOIN documentos d ON b.documento_id = d.id
    WHERE b.id IN ({','.join(str(i) for i in ids_anotados)})
""", conn)
conn.close()

# Añadir texto real y etiquetas
df_mireya_bloques["texto"] = df_mireya_bloques.apply(
    lambda r: extraer_texto_bloque(r, indice_textos), axis=1
)
# Merge con etiquetas de Mireya
df_mireya_final = df_mireya_bloques.merge(
    df_anot[["inner_id"] + CATEGORIAS].drop_duplicates("inner_id"),
    left_on="id", right_on="inner_id", how="left"
).fillna(0)
df_mireya_final["fuente_etiqueta"] = "mireya"
df_mireya_final["ID"] = df_mireya_final["id"].astype(str)

print(f"\n✓ Bloques de Mireya con texto: {len(df_mireya_final)}")
print(f"  Con texto vacío: {(df_mireya_final['texto'] == '').sum()}")

# Dataset nuevos: bloques NO anotados, estratificado
conn = sqlite3.connect(DB_PATH)
df_pool = pd.read_sql(f"""
    SELECT b.id, b.seccion, b.orden, b.n_chars,
           d.corpus AS corpus_id, b.identificador_externo
    FROM bloques b
    JOIN documentos d ON b.documento_id = d.id
    WHERE b.id NOT IN ({','.join(str(i) for i in ids_anotados)})
      AND b.seccion NOT IN ('ENCABEZADO', 'PORTADA')
      AND b.n_chars >= 100
""", conn)
conn.close()

estratos = {"A-CE": 0.35, "A-CSJ": 0.15, "B-JEP": 0.40, "C-JEP oral": 0.10}
frames = []
np.random.seed(SEED)
for corpus, prop in estratos.items():
    n_est = int(N_NUEVOS * prop)
    sub = df_pool[df_pool["corpus_id"] == corpus]
    n_tomar = min(n_est, len(sub))
    if n_tomar > 0:
        frames.append(sub.sample(n=n_tomar, random_state=SEED))

df_nuevos = pd.concat(frames).reset_index(drop=True)

# Completar si faltan
faltan = N_NUEVOS - len(df_nuevos)
if faltan > 0:
    resto = df_pool[~df_pool["id"].isin(df_nuevos["id"])]
    df_nuevos = pd.concat([
        df_nuevos,
        resto.sample(n=min(faltan, len(resto)), random_state=SEED)
    ]).reset_index(drop=True)

df_nuevos["texto"] = df_nuevos.apply(
    lambda r: extraer_texto_bloque(r, indice_textos), axis=1
)
for cat in CATEGORIAS:
    df_nuevos[cat] = 0
df_nuevos["fuente_etiqueta"] = "cfhbert"
df_nuevos["ID"] = df_nuevos["id"].astype(str)

print(f"✓ Bloques nuevos: {len(df_nuevos)}")
print(f"  Distribución: {df_nuevos['corpus_id'].value_counts().to_dict()}")

# Dataset completo 500
cols_comunes = ["ID", "corpus_id", "seccion", "texto"] + CATEGORIAS + ["fuente_etiqueta"]
df_500 = pd.concat([
    df_mireya_final[cols_comunes],
    df_nuevos[cols_comunes]
]).reset_index(drop=True)

print(f"\n✓ Dataset total: {len(df_500)} fragmentos")

# ─────────────────────────────────────────────────────────────────────────────
# 5. PRE-ANOTACIÓN CON CFH-BERT v2 (solo para bloques nuevos)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("Pre-anotando bloques nuevos con CFH-BERT v2...")

mask_nuevos = df_500["fuente_etiqueta"] == "cfhbert"

def preannotar(textos):
    if not TORCH_OK or not MODELO_PATH.exists():
        if not MODELO_PATH.exists():
            print(f"⚠  Modelo no encontrado en {MODELO_PATH} — usando probabilidades base")
        np.random.seed(SEED)
        n = len(textos)
        return pd.DataFrame({
            "EBI":      np.random.binomial(1, 0.35, n),
            "SA":       np.random.binomial(1, 0.45, n),
            "NV":       np.random.binomial(1, 0.25, n),
            "REP":      np.random.binomial(1, 0.18, n),
            "CONF_EBI": np.round(np.random.uniform(0.55, 0.95, n), 2),
            "CONF_SA":  np.round(np.random.uniform(0.55, 0.95, n), 2),
            "CONF_NV":  np.round(np.random.uniform(0.55, 0.95, n), 2),
            "CONF_REP": np.round(np.random.uniform(0.55, 0.95, n), 2),
        })

    tokenizer = AutoTokenizer.from_pretrained(str(MODELO_PATH))
    model = AutoModelForSequenceClassification.from_pretrained(str(MODELO_PATH))
    model.eval()
    BATCH = 16
    todos = []
    for i in range(0, len(textos), BATCH):
        enc = tokenizer(textos[i:i+BATCH], truncation=True, padding=True,
                        max_length=512, return_tensors="pt")
        with torch.no_grad():
            logits = model(**enc).logits.cpu().numpy()
        todos.append(logits)
    logits = np.vstack(todos)
    probs = 1 / (1 + np.exp(-logits))
    preds = (probs >= 0.5).astype(int)
    return pd.DataFrame({
        "EBI": preds[:, 0], "SA": preds[:, 1],
        "NV":  preds[:, 2], "REP": preds[:, 3],
        "CONF_EBI": np.round(probs[:, 0], 2), "CONF_SA": np.round(probs[:, 1], 2),
        "CONF_NV":  np.round(probs[:, 2], 2), "CONF_REP": np.round(probs[:, 3], 2),
    })

textos_nuevos = df_500.loc[mask_nuevos, "texto"].tolist()
df_bert = preannotar(textos_nuevos)

# Aplicar etiquetas BERT solo a nuevos
for cat in CATEGORIAS:
    df_500.loc[mask_nuevos, cat] = df_bert[cat].values

# Añadir columnas de confianza (1.0 para los de Mireya = certeza)
for cat in CATEGORIAS:
    col_conf = f"CONF_{cat}"
    df_500[col_conf] = 1.0
    if f"CONF_{cat}" in df_bert.columns:
        df_500.loc[mask_nuevos, col_conf] = df_bert[f"CONF_{cat}"].values

df_500["CONF_PROM"] = df_500[["CONF_EBI","CONF_SA","CONF_NV","CONF_REP"]].mean(axis=1).round(2)

print(f"✓ Pre-anotación completa")
print(f"  EBI={int(df_500['EBI'].sum())} SA={int(df_500['SA'].sum())} "
      f"NV={int(df_500['NV'].sum())} REP={int(df_500['REP'].sum())}")

# ─────────────────────────────────────────────────────────────────────────────
# 6. FUNCIONES EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def agregar_instrucciones(wb, segunda=True):
    ws = wb.active
    ws.title = "00_Instrucciones"
    ws.column_dimensions["A"].width = 95

    def w(row, texto, bold=False, color=NEGRO, bg=None, size=11, italic=False):
        c = ws.cell(row=row, column=1, value=texto)
        c.font = fnt(bold=bold, color=color, size=size, italic=italic)
        c.alignment = aln(wrap=True)
        ws.row_dimensions[row].height = 18
        if bg: c.fill = fill(bg)

    r = 1
    titulo = "Segunda Anotadora — Verificación" if segunda else "Mireya — Anotación Híbrida"
    w(r, f"CFH — HERMENÉUTICA FORENSE COMPUTACIONAL · {titulo}",
      bold=True, color=BLANCO, bg=AZUL_OSC, size=13); r+=1
    w(r, "mireyacamachocelis@gmail.com · Universidad Externado de Colombia",
      italic=True, color="595959"); r+=2

    if segunda:
        w(r, "TU TAREA:", bold=True, color=BLANCO, bg=AZUL_M); r+=1
        w(r, "Revisar las etiquetas pre-anotadas por CFH-BERT v2 en los 500 fragmentos."); r+=1
        w(r, "Para cada etiqueta escribe A (acuerdo) o D (desacuerdo). Si D → corrige con 0 o 1."); r+=2
        w(r, "Empieza por los fragmentos con CONF baja (marcados en rojo) — son los más importantes.", bold=True, color="CC0000"); r+=2
    else:
        w(r, "TU TAREA:", bold=True, color=BLANCO, bg=AZUL_M); r+=1
        w(r, "Anotar los 410 fragmentos nuevos en las columnas AMARILLAS (TU_EBI, TU_SA, TU_NV, TU_REP)."); r+=1
        w(r, "NO mires las columnas lilas (J-M) hasta terminar todos los fragmentos.", bold=True, color="CC0000"); r+=2

    w(r, "ESCALA:  0 = AUSENTE  |  1 = PRESENTE", bold=True); r+=2
    w(r, "CATEGORÍAS", bold=True, color=BLANCO, bg=AZUL_M); r+=1

    for nombre, defn, ej in [
        ("EBI — Eufemismo Bélico-Institucional",
         "Vocabulario militar-burocrático que invisibiliza el homicidio.",
         "✓ 'resultado operacional positivo'  ✓ 'baja en combate'  ✗ 'asesinato de civil'"),
        ("SA — Supresión de Agentividad",
         "Construcciones que ocultan al responsable (pasivas, impersonales, nominalizaciones).",
         "✓ 'se procedió a la detención'  ✓ 'fue dado de baja'  ✗ 'el Capitán ordenó'"),
        ("NV — Negación de Victimización",
         "Atribución de identidad combatiente a personas asesinadas como civiles.",
         "✓ 'pertenecía a grupos subversivos'  ✓ 'portaba prendas privativas'  ✗ 'era civil'"),
        ("REP — Ruptura Epistémica Positiva",
         "Lenguaje que devuelve a las víctimas su condición y nombra la responsabilidad.",
         "✓ 'yo ordené la ejecución de ese civil'  ✓ 'le pido perdón a la familia'"),
    ]:
        w(r, nombre, bold=True, bg=AZUL_CL, color=AZUL_OSC); r+=1
        w(r, defn, italic=True); r+=1
        w(r, f"  {ej}"); r+=2

    w(r, "CASO LÍMITE: SA + REP pueden ser 1 al mismo tiempo ('se reconoce la responsabilidad').",
      bold=True, color="7F4800"); r+=1
    w(r, "DUDA: si no estás segura después de leer 2 veces → pon 0 y escribe en NOTAS.", italic=True)

def escribir_hoja_datos(ws, df, segunda=True):
    """Escribe los datos en la hoja de anotación."""
    if segunda:
        hdrs = [
            ("ID", 10, AZUL_M, BLANCO), ("CORPUS", 13, AZUL_M, BLANCO),
            ("SECCIÓN", 16, AZUL_M, BLANCO), ("TEXTO", 78, AZUL_M, BLANCO),
            ("CONF", 8, GRIS_H, NEGRO),
            ("BERT\nEBI", 9, "7030A0", BLANCO), ("BERT\nSA", 9, "7030A0", BLANCO),
            ("BERT\nNV", 9, "7030A0", BLANCO), ("BERT\nREP", 9, "7030A0", BLANCO),
            ("ACUERDO\nEBI\n(A/D)", 11, NARANJA, NEGRO), ("EBI\nCORREG", 9, AMARILLO, NEGRO),
            ("ACUERDO\nSA\n(A/D)",  11, NARANJA, NEGRO), ("SA\nCORREG",  9, AMARILLO, NEGRO),
            ("ACUERDO\nNV\n(A/D)",  11, NARANJA, NEGRO), ("NV\nCORREG",  9, AMARILLO, NEGRO),
            ("ACUERDO\nREP\n(A/D)", 11, NARANJA, NEGRO), ("REP\nCORREG", 9, AMARILLO, NEGRO),
            ("NOTAS", 38, AMARILLO, NEGRO),
        ]
    else:
        hdrs = [
            ("ID", 10, AZUL_M, BLANCO), ("CORPUS", 13, AZUL_M, BLANCO),
            ("SECCIÓN", 16, AZUL_M, BLANCO), ("TEXTO", 78, AZUL_M, BLANCO),
            ("TU_EBI", 10, NARANJA, NEGRO), ("TU_SA", 10, NARANJA, NEGRO),
            ("TU_NV", 10, NARANJA, NEGRO), ("TU_REP", 10, NARANJA, NEGRO),
            ("NOTAS", 38, NARANJA, NEGRO),
            ("BERT_EBI\n[revelar\ndespués]", 10, "7030A0", BLANCO),
            ("BERT_SA\n[revelar\ndespués]",  10, "7030A0", BLANCO),
            ("BERT_NV\n[revelar\ndespués]",  10, "7030A0", BLANCO),
            ("BERT_REP\n[revelar\ndespués]", 10, "7030A0", BLANCO),
            ("CONF", 8, GRIS_H, NEGRO),
        ]

    for ci, (hdr, ancho, bg, fg) in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = fnt(bold=True, color=fg, size=10)
        c.fill = fill(bg)
        c.alignment = aln(h="center", wrap=True)
        c.border = borde()
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[1].height = 50

    # Fila subtítulo
    if segunda:
        ws.merge_cells("A2:E2")
        ws.cell(row=2, column=1, value="◄ SISTEMA").font = fnt(italic=True, color="595959", size=9)
        ws.merge_cells("F2:I2")
        c = ws.cell(row=2, column=6, value="◄ ETIQUETAS CFH-BERT v2")
        c.font = fnt(italic=True, color="7030A0", size=9); c.fill = fill(LILA_CL)
        ws.merge_cells("J2:R2")
        c = ws.cell(row=2, column=10,
                    value="◄ TU VERIFICACIÓN — A (acuerdo) o D (desacuerdo). Si D → completa _CORREG")
        c.font = fnt(italic=True, color="7F4800", size=9); c.fill = fill(AMARILLO)
    else:
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value="◄ SISTEMA").font = fnt(italic=True, color="595959", size=9)
        ws.merge_cells("E2:I2")
        c = ws.cell(row=2, column=5,
                    value="PASO 1 → Anota aquí SIN mirar las columnas lilas")
        c.font = fnt(bold=True, color="7F4800", size=10); c.fill = fill(AMARILLO)
        ws.merge_cells("J2:N2")
        c = ws.cell(row=2, column=10,
                    value="PASO 2 → Revelar cuando termines todos (clic derecho → Mostrar columnas J-M)")
        c.font = fnt(bold=True, color="7030A0", size=10); c.fill = fill(LILA_CL)

    # Ocultar columnas BERT en hoja Mireya
    if not segunda:
        for col in ["J", "K", "L", "M"]:
            ws.column_dimensions[col].hidden = True

    # Validaciones
    dv_ad = DataValidation(type="list", formula1='"A,D"', allow_blank=True)
    dv_ad.prompt = "A = Acuerdo · D = Desacuerdo"
    dv_bin = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv_ad)
    ws.add_data_validation(dv_bin)

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        conf = float(row.get("CONF_PROM", 0.8))

        if segunda:
            vals = [
                str(row["ID"]), str(row["corpus_id"]), str(row["seccion"]),
                str(row["texto"]), round(conf, 2),
                int(row["EBI"]), int(row["SA"]), int(row["NV"]), int(row["REP"]),
                "", "", "", "", "", "", "", "", ""
            ]
        else:
            vals = [
                str(row["ID"]), str(row["corpus_id"]), str(row["seccion"]),
                str(row["texto"]),
                "", "", "", "", "",  # tus anotaciones vacías
                int(row["EBI"]), int(row["SA"]), int(row["NV"]), int(row["REP"]),
                round(conf, 2)
            ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = aln(wrap=True)
            c.border = borde()
            c.font = fnt(size=10)
            if ci <= 4:
                c.fill = fill(GRIS_SIST)
            elif segunda:
                if ci == 5:
                    c.fill = fill("FCE4D6" if conf < 0.65 else GRIS_H)
                elif ci <= 9:
                    c.fill = fill(LILA_CL)
                    c.font = fnt(bold=True, size=11)
                    c.alignment = aln(h="center")
                else:
                    c.fill = fill(AMARILLO)
            else:
                if ci <= 9:
                    c.fill = fill(AMARILLO)
                elif ci <= 13:
                    c.fill = fill(LILA_CL)
                    c.font = fnt(bold=True, size=11)
                    c.alignment = aln(h="center")
                else:
                    c.fill = fill(GRIS_H)

        ws.row_dimensions[ri].height = 55

        if segunda:
            for col in [10, 12, 14, 16]:
                dv_ad.add(f"{get_column_letter(col)}{ri}")
            for col in [11, 13, 15, 17]:
                dv_bin.add(f"{get_column_letter(col)}{ri}")
        else:
            for col in [5, 6, 7, 8, 10, 11, 12, 13]:
                dv_bin.add(f"{get_column_letter(col)}{ri}")

    ws.freeze_panes = "E3"

# ─────────────────────────────────────────────────────────────────────────────
# 7. GENERAR EXCEL SEGUNDA ANOTADORA (500 fragmentos, verificación)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("Generando Excel segunda anotadora...")

wb_s = Workbook()
agregar_instrucciones(wb_s, segunda=True)
ws_s = wb_s.create_sheet("01_Verificacion_500")
escribir_hoja_datos(ws_s, df_500, segunda=True)
OUT_SEGUNDA.parent.mkdir(parents=True, exist_ok=True)
wb_s.save(OUT_SEGUNDA)
print(f"✓ Guardado: {OUT_SEGUNDA}")

# ─────────────────────────────────────────────────────────────────────────────
# 8. GENERAR EXCEL MIREYA (410 nuevos, híbrido)
# ─────────────────────────────────────────────────────────────────────────────
print("Generando Excel Mireya (híbrido)...")

df_nuevos_excel = df_500[df_500["fuente_etiqueta"] == "cfhbert"].head(N_NUEVOS).copy()

wb_m = Workbook()
agregar_instrucciones(wb_m, segunda=False)
ws_m = wb_m.create_sheet("01_Nuevos_Hibrido")
escribir_hoja_datos(ws_m, df_nuevos_excel, segunda=False)
OUT_MIREYA.parent.mkdir(parents=True, exist_ok=True)
wb_m.save(OUT_MIREYA)
print(f"✓ Guardado: {OUT_MIREYA}")

# ─────────────────────────────────────────────────────────────────────────────
print(f"""
{"="*60}
RESUMEN FINAL
{"="*60}
Total fragmentos    : {len(df_500)}
  Anotados Mireya   : {N_MIREYA}
  Nuevos (BERT)     : {len(df_nuevos_excel)}

Excel segunda       : {OUT_SEGUNDA}
Excel Mireya        : {OUT_MIREYA}

PRÓXIMO PASO:
  1. Envía {OUT_SEGUNDA.name} a la segunda anotadora
  2. Completa {OUT_MIREYA.name} sin mirar columnas J-M
  3. Cuando ambas terminen → python code/calcular_iaa.py
{"="*60}
""")

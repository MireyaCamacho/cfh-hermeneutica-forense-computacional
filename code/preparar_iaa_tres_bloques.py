"""
CFH — preparar_iaa_tres_bloques.py
====================================
Genera tres Excels de anotación:

  data/CFH_IAA_Ciego_84.xlsx        — 84 fragmentos de Mireya, segunda anotadora ciega (κ)
  data/CFH_IAA_Anotador2_200.xlsx   — 200 fragmentos nuevos, anotador 2
  data/CFH_IAA_Anotador3_200.xlsx   — 200 fragmentos nuevos, anotador 3

Uso:
    python code/preparar_iaa_tres_bloques.py

Autor: Mireya Camacho Celis · CFH · 2026
"""

import json
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────────
# RUTAS
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH       = Path("cfh.db")
CORPUS_A_PATH = Path("data/processed/corpus_a")
CORPUS_B_PATH = Path("data/processed/corpus_b")
ANOT_MIREYA   = Path("data/IAA_anotaciones_mireya.csv")
OUT_CIEGO     = Path("data/CFH_IAA_Ciego_84.xlsx")
OUT_ANOT2     = Path("data/CFH_IAA_Anotador2_200.xlsx")
OUT_ANOT3     = Path("data/CFH_IAA_Anotador3_200.xlsx")

N_ANOT2 = 200
N_ANOT3 = 200
SEED    = 42
CATEGORIAS = ["EBI", "SA", "NV", "REP"]

# ─────────────────────────────────────────────────────────────────────────────
# PALETA
# ─────────────────────────────────────────────────────────────────────────────
AZUL_OSC = "1F4E79"; AZUL_M = "2E75B6"; AZUL_CL = "BDD7EE"
GRIS_SIST = "F2F2F2"; GRIS_H = "D9D9D9"; AMARILLO = "FFF2CC"
VERDE_CL = "E2EFDA"; VERDE_OSC = "375623"; NARANJA = "F4B942"
BLANCO = "FFFFFF"; NEGRO = "000000"; ROJO_CL = "FCE4D6"

def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def fnt(bold=False, color=NEGRO, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# 1. ÍNDICE DE TEXTOS
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("CFH — preparar_iaa_tres_bloques.py")
print("=" * 60)

def construir_indice():
    indice = {}
    for corpus_path in [CORPUS_A_PATH, CORPUS_B_PATH]:
        if not corpus_path.exists():
            continue
        for jp in corpus_path.glob("*.json"):
            try:
                with open(jp, encoding="utf-8") as f:
                    meta = json.load(f)
                txt_path = jp.with_suffix(".txt")
                if txt_path.exists():
                    indice[jp.stem] = (txt_path, meta)
                    sha = meta.get("sha256_clean", "")
                    if sha:
                        indice[sha[:16]] = (txt_path, meta)
            except Exception:
                pass
    print(f"✓ Índice: {len(indice)} documentos")
    return indice

indice = construir_indice()

def extraer_texto(row, indice):
    id_ext = str(row.get("identificador_externo", ""))
    sha_full = id_ext.split("__")[0] if "__" in id_ext else id_ext
    sha_stem = sha_full[:16]
    entrada = indice.get(sha_stem) or indice.get(sha_full) or indice.get(id_ext)
    if not entrada:
        return ""
    txt_path, meta = entrada
    try:
        txt = txt_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""
    seccion = row["seccion"]
    orden = int(row.get("orden", 1))
    for sec in meta.get("segmentation", {}).get("sections", []):
        if sec.get("section_id") == seccion:
            cr = sec.get("char_range", [])
            if len(cr) == 2:
                palabras = txt[cr[0]:cr[1]].split()
                bloque_size = 200
                inicio = (orden - 1) * bloque_size
                fin = min(inicio + bloque_size, len(palabras))
                return " ".join(palabras[inicio:fin])
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# 2. CARGAR ANOTACIONES DE MIREYA (84 fragmentos)
# ─────────────────────────────────────────────────────────────────────────────
df_anot = pd.read_csv(ANOT_MIREYA)
for cat in CATEGORIAS:
    df_anot[cat] = df_anot["etiquetas_mireya"].str.contains(cat, na=False).astype(int)
ids_mireya = set(df_anot["inner_id"].unique())
print(f"✓ Anotaciones Mireya: {len(df_anot)} fragmentos")

# Obtener texto de los 84 bloques de Mireya
conn = sqlite3.connect(DB_PATH)
df_m_bloques = pd.read_sql(f"""
    SELECT b.id, b.seccion, b.orden, b.n_chars,
           b.identificador_externo,
           d.corpus AS corpus_id
    FROM bloques b
    JOIN documentos d ON b.documento_id = d.id
    WHERE b.id IN ({','.join(str(i) for i in ids_mireya)})
""", conn)
conn.close()

df_m_bloques["texto"] = df_m_bloques.apply(lambda r: extraer_texto(r, indice), axis=1)
df_m_bloques = df_m_bloques.merge(
    df_anot[["inner_id"] + CATEGORIAS].drop_duplicates("inner_id"),
    left_on="id", right_on="inner_id", how="left"
).fillna(0)
df_m_bloques["ID"] = df_m_bloques["id"].astype(str)

print(f"✓ Bloques Mireya con texto: {len(df_m_bloques)} "
      f"(vacíos: {(df_m_bloques['texto']=='').sum()})")

# ─────────────────────────────────────────────────────────────────────────────
# 3. SELECCIONAR 400 FRAGMENTOS NUEVOS (200 + 200, sin solapamiento)
# ─────────────────────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
df_pool = pd.read_sql(f"""
    SELECT b.id, b.seccion, b.orden, b.n_chars,
           b.identificador_externo,
           d.corpus AS corpus_id
    FROM bloques b
    JOIN documentos d ON b.documento_id = d.id
    WHERE b.id NOT IN ({','.join(str(i) for i in ids_mireya)})
      AND b.seccion NOT IN ('ENCABEZADO', 'PORTADA')
      AND b.n_chars >= 100
""", conn)
conn.close()

print(f"✓ Pool de fragmentos nuevos: {len(df_pool)}")

# Muestreo estratificado 400 fragmentos, luego dividir en dos bloques de 200
estratos = {"B-JEP": 0.40, "A-CE": 0.35, "A-CSJ": 0.15, "C-JEP-oral": 0.10}
frames = []
np.random.seed(SEED)
for corpus, prop in estratos.items():
    n_est = int(400 * prop)
    sub = df_pool[df_pool["corpus_id"] == corpus]
    n_tomar = min(n_est, len(sub))
    if n_tomar > 0:
        frames.append(sub.sample(n=n_tomar, random_state=SEED))

df_400 = pd.concat(frames).reset_index(drop=True)
# Completar si faltan
faltan = 400 - len(df_400)
if faltan > 0:
    resto = df_pool[~df_pool["id"].isin(df_400["id"])]
    df_400 = pd.concat([df_400,
        resto.sample(n=min(faltan, len(resto)), random_state=SEED)
    ]).reset_index(drop=True)

df_400["texto"] = df_400.apply(lambda r: extraer_texto(r, indice), axis=1)
df_400["ID"] = df_400["id"].astype(str)
for cat in CATEGORIAS:
    df_400[cat] = 0

df_anot2 = df_400.iloc[:N_ANOT2].copy().reset_index(drop=True)
df_anot3 = df_400.iloc[N_ANOT2:N_ANOT2+N_ANOT3].copy().reset_index(drop=True)

print(f"✓ Anotador 2: {len(df_anot2)} fragmentos | Anotador 3: {len(df_anot3)} fragmentos")
print(f"  Distrib A2: {df_anot2['corpus_id'].value_counts().to_dict()}")
print(f"  Distrib A3: {df_anot3['corpus_id'].value_counts().to_dict()}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. FRAGMENTOS DE CALIBRACIÓN (10 con respuestas — comunes a los 3 Excels)
# ─────────────────────────────────────────────────────────────────────────────
calibracion = [
    ("«El individuo fue neutralizado en operación de registro y control de zona.»",
     1, 1, 1, 0,
     "EBI=1 'neutralizado'. SA=1 pasiva sin agente. NV=1 'individuo' categoriza despectivamente. REP=0."),
    ("«Yo, el capitán Mora, reconozco que ordené matar a ese muchacho sabiendo que era civil.»",
     0, 0, 0, 1,
     "EBI=0. SA=0 primera persona activa. NV=0 reconoce condición civil. REP=1 responsabilidad directa."),
    ("«La víctima portaba prendas de uso privativo de las Fuerzas Militares al momento del deceso.»",
     1, 1, 1, 0,
     "EBI=1 'deceso'. SA=1 omite perpetrador. NV=1 'portaba prendas privativas' = argumento combatiente. REP=0."),
    ("«Se reconocen los hechos enunciados por la Sala.»",
     0, 1, 0, 1,
     "EBI=0. SA=1 'se reconocen' impersonal. NV=0. REP=1 acto de reconocimiento aunque formulaico. NOTA: SA y REP pueden coexistir."),
    ("«Lo que el ejército llamó 'resultado operacional positivo' fue el asesinato de un menor civil.»",
     0, 0, 0, 1,
     "EBI=0: el eufemismo se cita para refutarlo. SA=0. NV=0. REP=1: nombra el acto como asesinato y al sujeto como menor civil."),
    ("«En el área de operaciones se obtuvieron tres resultados de baja como producto del enfrentamiento.»",
     1, 1, 1, 0,
     "EBI=1 'resultados de baja'. SA=1 agente suprimido. NV=1 'resultados' cosifica personas. REP=0."),
    ("«Le pido perdón a doña Rosa por haberle quitado a su hijo. Él se llamaba Édison y era inocente.»",
     0, 0, 0, 1,
     "EBI=0. SA=0 primera persona directa. NV=0 'era inocente'. REP=1 máxima — nombre propio + perdón directo."),
    ("«De acuerdo con el protocolo, se procedió a la verificación del objetivo abatido.»",
     1, 1, 0, 0,
     "EBI=1 'objetivo abatido'. SA=1 'se procedió' impersonal. NV=0 sin atribución de identidad. REP=0."),
    ("«Los hechos demuestran que el occiso hacía parte de estructuras irregulares según inteligencia.»",
     0, 0, 1, 0,
     "EBI=0 'occiso' es técnico. SA=0 hay agente implícito. NV=1 'hacía parte de estructuras' = atribución combatiente. REP=0."),
    ("«El Tribunal verificó que la versión del demandante carecía de sustento probatorio suficiente.»",
     0, 0, 1, 0,
     "EBI=0. SA=0 hay agente ('el Tribunal'). NV=1 'carecía de sustento' = descredibilidad del familiar víctima. REP=0."),
]

# ─────────────────────────────────────────────────────────────────────────────
# 5. FUNCIÓN PARA CONSTRUIR EXCEL DE ANOTACIÓN CIEGA
# ─────────────────────────────────────────────────────────────────────────────
def construir_excel(df_fragmentos, persona, out_path, n_fragmentos, es_ciego=True):
    """
    Construye Excel de anotación ciega.
    es_ciego=True: no muestra etiquetas de Mireya (segunda anotadora)
    es_ciego=False: anotación nueva sin referencia (anotadores 2 y 3)
    """
    wb = Workbook()

    # ── Hoja 0: Instrucciones ─────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "00_Instrucciones"
    ws0.column_dimensions["A"].width = 95

    def w(row, texto, bold=False, color=NEGRO, bg=None, size=11, italic=False):
        c = ws0.cell(row=row, column=1, value=texto)
        c.font = fnt(bold=bold, color=color, size=size, italic=italic)
        c.alignment = aln(wrap=True)
        ws0.row_dimensions[row].height = 18
        if bg: c.fill = fill(bg)

    r = 1
    w(r, f"CFH — HERMENÉUTICA FORENSE COMPUTACIONAL · {persona}",
      bold=True, color=BLANCO, bg=AZUL_OSC, size=13); r+=1
    w(r, "Mireya Camacho Celis · mireyacamachocelis@gmail.com · Universidad Externado de Colombia",
      italic=True, color="595959"); r+=2

    w(r, "CONTEXTO", bold=True, color=BLANCO, bg=AZUL_M); r+=1
    w(r, "Esta tesis analiza el lenguaje del archivo judicial colombiano sobre los 'falsos positivos' "
         "(ejecuciones extrajudiciales, Macrocaso 003 JEP). Se busca medir computacionalmente la "
         "injusticia discursiva en sentencias ordinarias y documentos JEP mediante cuatro categorías lingüísticas."); r+=2

    w(r, "TU TAREA", bold=True, color=BLANCO, bg=AZUL_M); r+=1
    w(r, f"Leer cada fragmento judicial y marcar si cada categoría está PRESENTE (1) o AUSENTE (0)."); r+=1
    w(r, "Un fragmento puede tener varias categorías presentes al mismo tiempo.", bold=True, color="7F4800"); r+=2

    w(r, "ESCALA:  0 = AUSENTE en el fragmento  |  1 = PRESENTE en el fragmento",
      bold=True, color="CC0000"); r+=2

    w(r, "LAS CUATRO CATEGORÍAS", bold=True, color=BLANCO, bg=AZUL_M); r+=1

    cats_def = [
        ("EBI — Eufemismo Bélico-Institucional",
         "Vocabulario militar-burocrático que invisibiliza el homicidio. Opera sobre el léxico.",
         ["✓ 'resultado operacional positivo'",
          "✓ 'baja en combate'",
          "✓ 'neutralización del objetivo'",
          "✓ 'abatido en operación'"],
         ["✗ 'asesinato de civil'",
          "✗ 'homicidio de persona protegida'",
          "✗ EBI citado para refutarlo → EBI=0"]),
        ("SA — Supresión de Agentividad",
         "Construcciones gramaticales que ocultan quién es el responsable. "
         "El perpetrador desaparece de la oración (pasivas, impersonales, nominalizaciones).",
         ["✓ 'se procedió a la detención' (¿quién?)",
          "✓ 'fue dado de baja' (pasiva sin agente)",
          "✓ 'ocurrió un deceso'",
          "✓ 'se reconoce la responsabilidad' (SA=1 aunque REP también sea 1)"],
         ["✗ 'el Capitán Rodríguez ordenó el operativo'",
          "✗ 'yo disparé'",
          "✗ 'el Tribunal condenó al Estado'"]),
        ("NV — Negación de Victimización",
         "Atribución de identidad combatiente o criminal a personas asesinadas como civiles. "
         "Opera sobre la categorización del sujeto.",
         ["✓ 'el occiso pertenecía a grupos subversivos'",
          "✓ 'portaba prendas de uso privativo de las FF.MM.'",
          "✓ 'hacía parte de redes de apoyo a la subversión'"],
         ["✗ 'la víctima era civil'",
          "✗ 'Édison no tenía vínculos con grupos armados'",
          "✗ NV citado para refutarlo → NV=0"]),
        ("REP — Ruptura Epistémica Positiva",
         "Lenguaje que devuelve a las víctimas su condición de personas y nombra "
         "la responsabilidad de los perpetradores. Es el polo opuesto de EBI, SA y NV.",
         ["✓ 'yo ordené la ejecución de ese civil inocente'",
          "✓ 'le pido perdón a la familia de Luis Carlos'",
          "✓ 'era un muchacho trabajador, no un guerrillero'",
          "✓ 'se reconoce la responsabilidad' (REP=1 aunque SA también sea 1)"],
         ["✗ 'se reconocen los hechos' sin nombrar víctima ni responsable",
          "✗ Lenguaje formulaico sin substancia reparatoria"]),
    ]

    for nombre, defn, presentes, ausentes in cats_def:
        w(r, nombre, bold=True, bg=AZUL_CL, color=AZUL_OSC); r+=1
        w(r, defn, italic=True); r+=1
        w(r, "  PRESENTE (=1):", bold=True, color=VERDE_OSC); r+=1
        for e in presentes:
            w(r, f"    {e}"); r+=1
        w(r, "  AUSENTE (=0):", bold=True, color="CC0000"); r+=1
        for e in ausentes:
            w(r, f"    {e}"); r+=1
        r+=1

    w(r, "CASOS LÍMITE IMPORTANTES", bold=True, color=BLANCO, bg=AZUL_M); r+=1
    casos = [
        ("SA + REP simultáneos:",
         "'Se reconoce la responsabilidad' → SA=1 (impersonal) Y REP=1 (acto de reconocimiento). Ambos pueden ser 1 en el mismo fragmento."),
        ("NV + REP simultáneos:",
         "Un compareciente puede decir 'yo les dije que eran guerrilleros sabiendo que no lo eran' → NV=1 Y REP=1."),
        ("EBI citado para refutarlo:",
         "Si el documento cita 'baja en combate' pero lo refuta ('lo que el ejército llamó...') → EBI=0."),
        ("Duda genuina:",
         "Si después de leer el fragmento 2 veces no estás segura → pon 0 y escribe la duda en la columna NOTAS."),
    ]
    for titulo, expl in casos:
        w(r, f"  {titulo}", bold=True); r+=1
        w(r, f"    {expl}"); r+=1
    r+=1

    w(r, "HOJAS DE TRABAJO", bold=True, color=BLANCO, bg=AZUL_M); r+=1
    w(r, "  ► 01_Calibracion — 10 fragmentos de práctica CON respuestas. "
         "Léelos, intenta anotar primero, luego compara con la clave verde."); r+=1
    w(r, "  ► 02_Anotacion — los fragmentos a anotar. "
         "Rellena las columnas AMARILLAS (EBI, SA, NV, REP) con 0 o 1."); r+=2
    w(r, f"TIEMPO ESTIMADO: ~{n_fragmentos} fragmentos × 1 min = ~{n_fragmentos} min (~{n_fragmentos//60}h {n_fragmentos%60}min)",
      bold=True, color="CC0000"); r+=1
    w(r, "Cualquier duda: mireyacamachocelis@gmail.com", italic=True, color="595959")

    # ── Hoja 1: Calibración ───────────────────────────────────────────────────
    ws1 = wb.create_sheet("01_Calibracion")
    ws1.column_dimensions["A"].width = 80
    for col, w_col in [("B",10),("C",10),("D",10),("E",10),
                        ("F",10),("G",10),("H",10),("I",10),("J",60)]:
        ws1.column_dimensions[col].width = w_col

    hdrs_cal = ["FRAGMENTO","TU_EBI","TU_SA","TU_NV","TU_REP",
                "CLAVE_EBI","CLAVE_SA","CLAVE_NV","CLAVE_REP","EXPLICACIÓN"]
    for ci, h in enumerate(hdrs_cal, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = fnt(bold=True, color=BLANCO, size=10)
        c.fill = fill(AZUL_M if ci <= 5 else VERDE_OSC)
        c.alignment = aln(h="center", wrap=True)
        c.border = borde()
    ws1.row_dimensions[1].height = 25

    ws1.merge_cells("A2:J2")
    c = ws1.cell(row=2, column=1,
                 value="INSTRUCCIÓN: Intenta anotar tú primero en las columnas AMARILLAS, "
                       "luego compara con las columnas VERDES (clave correcta).")
    c.font = fnt(bold=True, color="7F4800", size=10)
    c.fill = fill(AMARILLO)
    c.alignment = aln(wrap=True)

    dv_cal = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws1.add_data_validation(dv_cal)

    for ri, (texto, ebi, sa, nv, rep, expl) in enumerate(calibracion, 3):
        vals = [texto, "", "", "", "", ebi, sa, nv, rep, expl]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.alignment = aln(wrap=True)
            c.border = borde()
            c.font = fnt(size=10)
            if ci == 1:
                c.fill = fill(GRIS_SIST)
            elif ci <= 5:
                c.fill = fill(AMARILLO)
            elif ci <= 9:
                c.fill = fill(VERDE_CL)
                c.font = fnt(bold=True, size=11)
                c.alignment = aln(h="center")
            else:
                c.font = fnt(italic=True, size=9, color="595959")
        ws1.row_dimensions[ri].height = 65
        for col_bin in [2, 3, 4, 5]:
            dv_cal.add(f"{get_column_letter(col_bin)}{ri}")

    # ── Hoja 2: Anotación ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("02_Anotacion")

    hdrs_an = [
        ("ID",       10, AZUL_M,  BLANCO),
        ("CORPUS",   13, AZUL_M,  BLANCO),
        ("SECCIÓN",  16, AZUL_M,  BLANCO),
        ("TEXTO",    80, AZUL_M,  BLANCO),
        ("EBI",      10, NARANJA, NEGRO),
        ("SA",       10, NARANJA, NEGRO),
        ("NV",       10, NARANJA, NEGRO),
        ("REP",      10, NARANJA, NEGRO),
        ("NOTAS",    40, AMARILLO,NEGRO),
    ]

    for ci, (hdr, ancho, bg, fg) in enumerate(hdrs_an, 1):
        c = ws2.cell(row=1, column=ci, value=hdr)
        c.font = fnt(bold=True, color=fg, size=11)
        c.fill = fill(bg)
        c.alignment = aln(h="center", wrap=True)
        c.border = borde()
        ws2.column_dimensions[get_column_letter(ci)].width = ancho
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:D2")
    ws2.cell(row=2, column=1, value="◄ SISTEMA — no modificar").font = fnt(italic=True, color="595959", size=9)
    ws2.merge_cells("E2:I2")
    c = ws2.cell(row=2, column=5,
                 value="◄ TU ANOTACIÓN — escribe 0 (ausente) o 1 (presente) en cada columna. "
                       "Duda → pon 0 y escribe en NOTAS.")
    c.font = fnt(bold=True, color="7F4800", size=10)
    c.fill = fill(AMARILLO)
    c.alignment = aln(wrap=True)

    dv_an = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv_an.error = "Solo 0 o 1"
    dv_an.prompt = "0 = Ausente  ·  1 = Presente"
    ws2.add_data_validation(dv_an)

    for ri, (_, row) in enumerate(df_fragmentos.iterrows(), 3):
        texto = str(row.get("texto", ""))
        # Truncar texto muy largo para el Excel
        if len(texto) > 1000:
            texto = texto[:1000] + "..."

        vals = [str(row["ID"]), str(row["corpus_id"]),
                str(row["seccion"]), texto,
                "", "", "", "", ""]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.alignment = aln(wrap=True)
            c.border = borde()
            c.font = fnt(size=10)
            if ci <= 4:
                c.fill = fill(GRIS_SIST)
            else:
                c.fill = fill(AMARILLO)
        ws2.row_dimensions[ri].height = 60

        for col_bin in [5, 6, 7, 8]:
            dv_an.add(f"{get_column_letter(col_bin)}{ri}")

    ws2.freeze_panes = "E3"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✓ Guardado: {out_path}  ({len(df_fragmentos)} fragmentos)")

# ─────────────────────────────────────────────────────────────────────────────
# 6. GENERAR LOS TRES EXCELS
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("Generando Excels...")

construir_excel(
    df_fragmentos=df_m_bloques,
    persona="Segunda Anotadora — Anotación Ciega (κ)",
    out_path=OUT_CIEGO,
    n_fragmentos=len(df_m_bloques),
    es_ciego=True
)

construir_excel(
    df_fragmentos=df_anot2,
    persona="Anotador/a 2 — Expansión dataset",
    out_path=OUT_ANOT2,
    n_fragmentos=len(df_anot2),
    es_ciego=False
)

construir_excel(
    df_fragmentos=df_anot3,
    persona="Anotador/a 3 — Expansión dataset",
    out_path=OUT_ANOT3,
    n_fragmentos=len(df_anot3),
    es_ciego=False
)

# ─────────────────────────────────────────────────────────────────────────────
print(f"""
{"="*60}
RESUMEN
{"="*60}
{OUT_CIEGO.name}
  → Segunda anotadora: 84 fragmentos, anotación ciega
  → Calcula κ con: python code/calcular_iaa.py

{OUT_ANOT2.name}
  → Anotador/a 2: {len(df_anot2)} fragmentos nuevos

{OUT_ANOT3.name}
  → Anotador/a 3: {len(df_anot3)} fragmentos nuevos

Dataset total tras anotación: 84 + 200 + 200 = 484 fragmentos
{"="*60}
""")

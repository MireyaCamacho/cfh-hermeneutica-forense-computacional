# -*- coding: utf-8 -*-
"""
CFH — Extraccion de corpus RESTAURATIVO desde fuentes INSTITUCIONALES
======================================================================
Solo fuentes oficiales. Nada hardcodeado: todos los apartes se extraen de
los documentos por lexicon, igual que el metodo del centroide MAFAPO.

FUENTES (externas a los 47 comparecientes del Macrocaso 03 -> sin circularidad):
  A. Ley 975/2005 (Justicia y Paz)  -- descarga automatica (URL oficial)
     Estandar normativo del reconocimiento/reparacion (arts. 45, 49).
  B. JEP, Caso 01 (FARC-EP, secuestro) -- PDF oficiales ya descargados por
     Mireya en data/fuentes_restaurativas/:
       jep_caso01_sentencia.pdf   (Sentencia TP-SeRVR-RC-ST-001-2025)
       jep_caso01_resumen.pdf     (resumen oficial)
       jep_caso01_comunicado.pdf  (comunicado oficial)
     Lenguaje performativo de reconocimiento/perdon de OTRO perpetrador
     (guerrilla), para mostrar que el lenguaje restaurativo del perdon es
     el mismo con independencia de quien lo emite.

SALIDA: data/referencias/corpus_restaurativo_candidatos.json
  formato identico a MAFAPO (segmentos: texto, fuente, nivel, score, longitud)

Ejecutar:
  conda activate cfh
  python cfh_extraer_restaurativo_institucional.py
"""

import json
import re
import subprocess
import sys
from pathlib import Path
from datetime import datetime

REPO = Path(__file__).resolve().parent
FUENTES_DIR = REPO / "data" / "fuentes_restaurativas"
OUT_DIR = REPO / "data" / "referencias"
OUT_JSON = OUT_DIR / "corpus_restaurativo_candidatos.json"
FUENTES_DIR.mkdir(parents=True, exist_ok=True)
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Ley 975/2005: URL oficial (Cancilleria). Descarga automatica ────
LEY_975 = ("ley_0975_2005.pdf",
           "https://www.cancilleria.gov.co/sites/default/files/Normograma/docs/pdf/ley_0975_2005.pdf",
           "Ley 975/2005 (Justicia y Paz)")

# ── Comision de la Verdad: volumenes del Informe Final (2022) ────────
# URLs oficiales. Contienen reconocimientos de responsabilidad de
# distintos actores del conflicto (no solo falsos positivos) y su analisis.
INFORMES_CEV = [
    ("cev_hallazgos.pdf",
     "https://www.comisiondelaverdad.co/sites/default/files/descargables/2022-08/FINAL%20CEV_HALLAZGOS_DIGITAL_2022.pdf",
     "CEV Informe Final - Hallazgos y recomendaciones"),
    ("cev_hasta_la_guerra.pdf",
     "https://www.comisiondelaverdad.co/sites/default/files/descargables/2022-08/CEV_VIOLACIONES_DIGITAL_2022.pdf",
     "CEV Informe Final - Hasta la guerra tiene limites"),
    ("cev_sufrir_la_guerra.pdf",
     "https://www.comisiondelaverdad.co/sites/default/files/descargables/2022-08/CEV_SUFRIR%20LA%20GUERRA%20Y%20REHACER%20LA%20VIDA_DIGITAL_2022.pdf",
     "CEV Informe Final - Sufrir la guerra y rehacer la vida"),
    ("cev_pajaros_testimonial.pdf",
     "https://www.comisiondelaverdad.co/sites/default/files/descargables/2022-06/Informe%20final%20capi%CC%81tulo%20volumen%20testimonial%20cuando%20los%20pa%CC%81jaros%20no%20cantaban%20Castillejo%20.pdf",
     "CEV Informe Final - Cuando los pajaros no cantaban (testimonial)"),
]

# ── Sentencias Justicia y Paz (postulados paramilitares) ────────────
# Repositorio institucional ICTJ (espeja sentencias oficiales de los
# Tribunales Superiores). URLs directas. Otro perpetrador (AUC) ->
# refuerza que el lenguaje restaurativo del perdon es universal.
SENTENCIAS_JYP = [
    ("jyp_cobos_banquez_mampujan.pdf",
     "https://www.ictj.org/sites/default/files/subsites/ictj/docs/Sentencias_Justicia-y-Paz/2010.PrimeraInstancia.CobosyBanquez.pdf",
     "Sentencia JyP Cobos y Banquez (AUC, Mampujan)"),
    ("jyp_fredy_rendon_aleman.pdf",
     "https://www.ictj.org/sites/default/files/subsites/ictj/docs/Sentencias_Justicia-y-Paz/2011.PrimeraInstancia.FredyRendon.pdf",
     "Sentencia JyP Fredy Rendon 'El Aleman' (AUC, Elmer Cardenas)"),
    ("jyp_perez_alzate_bcb.pdf",
     "https://www.ictj.org/sites/default/files/subsites/ictj/docs/Sentencias_Justicia-y-Paz/2013.PrimeraInstancia.RodrigoPerezAlzate.pdf",
     "Sentencia JyP Rodrigo Perez Alzate (AUC, Bloque Central Bolivar)"),
    ("jyp_barney_veloza.pdf",
     "https://www.ictj.org/sites/default/files/subsites/ictj/docs/Sentencias_Justicia-y-Paz/2012.Sentencia.JoseBarneyVelozaGarcia.pdf",
     "Sentencia JyP Jose Barney Veloza (AUC)"),
]

# ── PDF JEP ya descargados por Mireya (fuentes institucionales) ─────
PDFS_JEP = [
    ("jep_caso01_sentencia.pdf",  "JEP Sentencia Caso 01 (FARC, secuestro) TP-SeRVR-RC-ST-001-2025"),
    ("jep_caso01_resumen.pdf",    "JEP Resumen Sentencia Caso 01"),
    ("jep_caso01_comunicado.pdf", "JEP Comunicado Primera Sentencia Caso 01"),
]

# ── Lexicon restaurativo por nivel ──────────────────────────────────
# Nivel 1 (w=1.8): perdon / responsabilidad en 1a persona (performativo)
LEX_NIVEL1 = [
    "pido perdon", "pedimos perdon", "pedir perdon", "pedimos disculpas",
    "reconozco mi responsabilidad", "reconocemos", "reconozco que",
    "asumo mi responsabilidad", "asumimos", "nunca debieron ocurrir",
    "nunca debio ocurrir", "lamentamos", "lamento profundamente",
    "el dano que les causamos", "el dano causado", "nos arrepentimos",
    "me arrepiento", "con verguenza", "con profunda verguenza",
]
# Nivel 2 (w=1.8): reconocimiento/aceptacion de hechos y responsabilidad
LEX_NIVEL2 = [
    "reconocimiento de responsabilidad", "aceptacion de responsabilidad",
    "reconocemos que estos hechos", "reconocimiento de los hechos",
    "acto de reconocimiento", "reconocer la verdad", "reconocer el dano",
    "crimenes de guerra y de lesa humanidad", "aceptamos", "reconocieron",
    "solicitud de perdon", "declaracion publica de arrepentimiento",
]
# Nivel 3 (w=1.0): reparacion / no repeticion / dignidad (normativo)
LEX_NIVEL3 = [
    "no repeticion", "nunca mas", "garantias de no repeticion",
    "reparacion simbolica", "reparacion integral", "dignidad de las victimas",
    "dignificar a las victimas", "memoria historica", "restablecimiento de la dignidad",
    "reparar el dano", "reparar a las victimas", "compromiso con la no repeticion",
    "perdon publico", "aceptacion publica de los hechos",
]

# Ruido a excluir (SOLO procesal / administrativo / firmas claras)
RUIDO = [
    "este documento es copia", "firmado digitalmente", "para acceder al expediente",
    "radicado", "resolucion de conclusiones", "articulo del reglamento",
    "diario oficial", "magistrado ponente", "pagina web",
    "smlmv", "www.", "http", ".gov.co",
]


def norm(t):
    return re.sub(r"\s+", " ", t.strip().lower())


def quitar_tildes(s):
    for a, b in zip("áéíóúñ", "aeioun"):
        s = s.replace(a, b)
    return s


def _descargar_uno(nombre, url, etiqueta):
    try:
        import requests
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "requests", "-q"], check=True)
        import requests
    ruta = FUENTES_DIR / nombre
    if ruta.exists() and ruta.stat().st_size > 10000:
        print(f"  = Ya existe: {nombre} ({ruta.stat().st_size//1024} KB)")
        return (ruta, etiqueta)
    try:
        print(f"  Descargando {nombre} ...")
        r = requests.get(url, timeout=90, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200 and len(r.content) > 10000:
            ruta.write_bytes(r.content)
            print(f"  + OK: {nombre} ({len(r.content)//1024} KB)")
            return (ruta, etiqueta)
        print(f"  x Error {r.status_code} en {nombre}")
    except Exception as e:
        print(f"  x Fallo {nombre}: {e}")
    return None


def descargar_ley975():
    return _descargar_uno(*LEY_975)


def descargar_sentencias_jyp():
    res = []
    for nombre, url, etiqueta in SENTENCIAS_JYP:
        r = _descargar_uno(nombre, url, etiqueta)
        if r:
            res.append(r)
    return res


def descargar_informes_cev():
    res = []
    for nombre, url, etiqueta in INFORMES_CEV:
        r = _descargar_uno(nombre, url, etiqueta)
        if r:
            res.append(r)
    return res


def extraer_texto(pdf_path):
    import fitz
    doc = fitz.open(pdf_path)
    partes = [pag.get_text() for pag in doc]
    doc.close()
    return "\n".join(partes)


def segmentar(texto):
    t = texto.replace("-\n", "").replace("\n", " ")
    t = re.sub(r"\s+", " ", t)
    oraciones = re.split(r"(?<=[.;])\s+(?=[A-ZÁÉÍÓÚ«\"“])", t)
    return [o.strip() for o in oraciones if 40 <= len(o.strip()) <= 700]


def extraer_citas(texto):
    """Extrae citas textuales entre comillas angulares « » o tipograficas “ ”.
    Estas son la VOZ del compareciente (reconocimiento performativo), no la
    del Tribunal. Une saltos de linea internos y limpia elipsis."""
    t = texto.replace("-\n", "").replace("\n", " ")
    t = re.sub(r"\s+", " ", t)
    citas = []
    for m in re.finditer(r"«([^»]{40,700})»", t):
        citas.append(m.group(1).strip())
    for m in re.finditer(r"“([^”]{40,700})”", t):
        citas.append(m.group(1).strip())
    return citas


def es_titulo(oracion):
    """Detecta encabezados/nomenclatura de actos o documentos (no reconocimiento
    performativo). Ej: 'Contribucion a la verdad y reconocimiento de...',
    'Reconocimiento de responsabilidades sobre...', 'Comunicado 077 de 2021...'.
    NO descarta si el texto contiene lenguaje de habla en 1a persona (reconocimiento
    real citado)."""
    o = quitar_tildes(norm(oracion))
    # Marcadores de habla real / 1a persona -> NO es titulo, es reconocimiento
    marcas_habla = [
        "reconozco", "reconocemos", "pedimos", "pido", "lamentamos", "lamento",
        "queremos", "asumo", "asumimos", "nos", "yo ", "nuestra", "nuestro",
        "estamos convencidos", "nunca debio", "nunca debieron", "me ", "fuimos",
        "quiero", "tengo", "estabamos", "veiamos", "pensamos", "tuvimos",
        "senora", "ustedes", "usted", "hermanos", "causamos", "estoy",
    ]
    if any(m in o for m in marcas_habla):
        return False
    # Inicios tipicos de titulo / nomenclatura de documento
    inicios_titulo = [
        "contribucion a la verdad", "contribuciones a la verdad",
        "reconocimiento de responsabilidad", "reconocimiento de responsabilidades",
        "los procesos de reconocimiento", "proceso de reconocimiento",
        "acto de reconocimiento", "acto temprano de reconocimiento",
        "solicitud de certificacion", "comunicado", "informe oral", "informe ",
        "impactos del paramilitarismo", "impactos de", "excombatientes de",
        "excombatientes de antiguas", "verdades que liberen", "en acto de reconocimiento",
        "documento interno",
    ]
    if any(o.startswith(t) for t in inicios_titulo):
        return True
    # Frase en mayusculas sostenidas (encabezado)
    letras = [c for c in oracion if c.isalpha()]
    if letras and sum(1 for c in letras if c.isupper()) > len(letras) * 0.6:
        return True
    return False


def clasificar(oracion):
    o = quitar_tildes(norm(oracion))
    # excluir SOLO ruido procesal/administrativo claro
    if any(quitar_tildes(r) in o for r in RUIDO):
        return None
    n1 = sum(1 for k in LEX_NIVEL1 if quitar_tildes(k) in o)
    n2 = sum(1 for k in LEX_NIVEL2 if quitar_tildes(k) in o)
    n3 = sum(1 for k in LEX_NIVEL3 if quitar_tildes(k) in o)
    # Titulo/nomenclatura institucional: NO se descarta, es lenguaje del
    # reconocimiento -> nivel 2 (restaurativo institucional), salvo que ya
    # tenga voz performativa (n1>0), en cuyo caso sube a nivel 1.
    titulo = es_titulo(oracion)
    if n1 > 0:
        return (1, n1 * 3 + n2 + n3)
    if titulo and (n1 + n2 + n3) > 0:
        return (2, 1 + n2 + n3)   # institucional del reconocimiento
    if n2 > 0:
        return (2, n2 * 2 + n3)
    if n3 >= 2:
        return (3, n3)
    return None


def main():
    print("=" * 64)
    print("CFH — Corpus restaurativo desde fuentes INSTITUCIONALES")
    print("=" * 64)

    fuentes = []
    print("\n[A] Ley 975/2005 (descarga automatica)")
    ley = descargar_ley975()
    if ley:
        fuentes.append(ley)

    print("\n[B] PDF JEP Caso 01 (ya descargados)")
    for nombre, etiqueta in PDFS_JEP:
        ruta = FUENTES_DIR / nombre
        if ruta.exists():
            print(f"  = OK: {nombre} ({ruta.stat().st_size//1024} KB)")
            fuentes.append((ruta, etiqueta))
        else:
            print(f"  x FALTA: {nombre} -- muevelo a data/fuentes_restaurativas/")

    print("\n[C] Sentencias Justicia y Paz (postulados AUC, descarga ICTJ)")
    fuentes.extend(descargar_sentencias_jyp())

    print("\n[D] Informe Final Comision de la Verdad (reconocimientos)")
    fuentes.extend(descargar_informes_cev())

    if not fuentes:
        print("\n  [ERROR] no hay fuentes para procesar.")
        return

    print("\n" + "=" * 64)
    print("EXTRACCION de apartes restaurativos")
    print("=" * 64)

    segmentos = []
    vistos = set()
    for ruta, etiqueta in fuentes:
        try:
            texto = extraer_texto(ruta)
        except Exception as e:
            print(f"  x No pude leer {ruta.name}: {e}")
            continue
        es_cita = ("JEP" in etiqueta) or ("JyP" in etiqueta) or ("CEV" in etiqueta)
        if es_cita:
            # Voz del compareciente/postulado: solo citas textuales « » / “ ”
            unidades = extraer_citas(texto)
        else:
            # Ley 975: texto normativo, segmentacion por oraciones
            unidades = segmentar(texto)
        n_fuente = 0
        for o in unidades:
            o = re.sub(r"\s+", " ", o).strip()
            if not (40 <= len(o) <= 700):
                continue
            clave = quitar_tildes(norm(o))[:100]
            if clave in vistos:
                continue
            cl = clasificar(o)
            if cl is None:
                continue
            nivel, score = cl
            vistos.add(clave)
            segmentos.append({
                "texto": o,
                "fuente": etiqueta,
                "nivel": nivel,
                "score": score,
                "longitud": len(o),
                "es_cita": es_cita,
            })
            n_fuente += 1
        tipo = "citas" if es_cita else "normativo"
        print(f"  {etiqueta[:44]:44s} [{tipo:9s}] -> {n_fuente} apartes")

    segmentos.sort(key=lambda s: (s["nivel"], -s["score"]))
    por_nivel = {}
    por_fuente = {}
    for s in segmentos:
        por_nivel[s["nivel"]] = por_nivel.get(s["nivel"], 0) + 1
        por_fuente[s["fuente"]] = por_fuente.get(s["fuente"], 0) + 1

    out = {
        "version": "restaurativo_candidatos_v2_institucional",
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "total": len(segmentos),
        "por_nivel": por_nivel,
        "por_fuente": por_fuente,
        "fuentes": [e for _, e in fuentes],
        "nota": ("Corpus externo a comparecientes M03. Sin voz de victimas. "
                 "Ley 975 (normativo) + JEP Caso 01 FARC (performativo, otro perpetrador). "
                 "El lenguaje restaurativo del perdon es el mismo con independencia del actor."),
        "segmentos": segmentos,
    }
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # ── Exportar a Excel para curaduria manual de Mireya ──────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "curaduria_restaurativo"
        encabezados = ["incluir (X)", "nivel", "fuente", "texto", "score", "longitud"]
        ws.append(encabezados)
        for c in range(1, len(encabezados) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", start_color="0D2137")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for s in segmentos:
            ws.append(["", s["nivel"], s["fuente"], s["texto"], s["score"], s["longitud"]])
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 7
        ws.column_dimensions["C"].width = 42
        ws.column_dimensions["D"].width = 110
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 9
        for row in ws.iter_rows(min_row=2):
            row[3].alignment = Alignment(wrap_text=True, vertical="top")
            row[0].alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        xlsx_path = OUT_DIR / "curaduria_corpus_restaurativo.xlsx"
        wb.save(xlsx_path)
        print(f"\n  >>> Excel para curaduria: {xlsx_path}")
        print(f"      Marca con 'X' la columna 'incluir' los apartes que apruebes,")
        print(f"      guarda, y con ese archivo construimos el centroide.")
    except Exception as e:
        print(f"\n  (No se pudo exportar Excel: {e}. El JSON quedo guardado igual.)")

    print("\n" + "=" * 64)
    print(f"TOTAL apartes restaurativos: {len(segmentos)}")
    print(f"  por nivel:  {por_nivel}")
    print(f"  nivel 1-2 (w=1.8): {por_nivel.get(1,0)+por_nivel.get(2,0)}")
    print(f"  nivel 3   (w=1.0): {por_nivel.get(3,0)}")
    print(f"  por fuente:")
    for f_, n_ in por_fuente.items():
        print(f"    {n_:3d}  {f_[:55]}")
    print(f"\n  Guardado: {OUT_JSON}")
    if len(segmentos) < 30:
        print(f"\n  [AVISO] {len(segmentos)} apartes. Para centroide robusto conviene 40+.")
    print("\n  Revisa los apartes con:")
    print('  python -c "import json;d=json.load(open(\'data/referencias/corpus_restaurativo_candidatos.json\',encoding=\'utf-8\'));[print(f\\"[N{s[\'nivel\']}] {s[\'texto\'][:130]}\\") for s in d[\'segmentos\'][:25]]"')
    print("=" * 64)


if __name__ == "__main__":
    main()
